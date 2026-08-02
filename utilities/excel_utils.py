from openpyxl import load_workbook


class ExcelUtils:

    @staticmethod
    def get_row_count(file, sheet):
        workbook = load_workbook(file)
        worksheet = workbook[sheet]
        return worksheet.max_row

    @staticmethod
    def read_data(file, sheet, row, column):
        workbook = load_workbook(file)
        worksheet = workbook[sheet]
        return worksheet.cell(row=row, column=column).value